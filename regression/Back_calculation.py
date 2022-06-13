# -*- coding: utf-8 -*-
"""
Created on Fri Jun 10 09:33:37 2022

@author: AYBR
"""

import PySimpleGUI as sg
import mysql.connector
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import math
import pandas as pd
from sklearn.linear_model import LinearRegression 
from math import *
import os, xlsxwriter
import array
import os
from os import walk
import math
import sys
import ctypes
from datetime import datetime, date, timezone
from collections import Counter
import xlsxwriter
import openpyxl
import shutil

#Main function
class Main:  

    print('============================Start=================================')
    plt.close('all')

def __init__(self): 
        self
# =============================================================================
#Read of MYSQL database - Parameters Level 1
# =============================================================================
conn=mysql.connector.connect(user='owdb_user',password='ituotdowdb',host='dklycopilod1',database='owdb')
mycursor=conn.cursor()
mycursor.execute("""SELECT * FROM owdb.Pisa_Param_Level_1""")
db_all = mycursor.fetchall()
df=db_all
    
df1=pd.DataFrame(df, columns=['Project','Soil_unit','Parameter','Type_f','Dep_f','Type_g','Dep_g','Rev','Coe_1_1','Coe_1_2','Coe_1_3','Coe_2_1','Coe_2_2','Coe_2_3','Coe_3_1','Coe_3_2','Coe_3_3'])
print('==================================================================')
print('Dataframe extracted from mySQL:')
print(df1)
conn.close()

# =============================================================================
#Read of MYSQL database - All design profiles
# =============================================================================
conn=mysql.connector.connect(user='ewdb_user',password='ituotdewdb',host='dklycopilod1',database='ewdb')
mycursor=conn.cursor()
mycursor.execute("""SELECT * FROM ewdb.soil_data_source""")
db_all = mycursor.fetchall()
df=db_all
    

df2=pd.DataFrame(df, columns=['id','layer','soil_revision','Depth','Unit','w','Gam','Ip','FC','St','eps50','qt_LE','qt_CLBE','qt_BE','qt_CHBE','qt_HE','fs_LE','fs_CLBE','fs_BE','fs_CHBE','fs_HE','su_LE','su_CLBE','su_BE','su_CHBE','su_HE','Dr','phi_LE','phi_CLBE','phi_BE','phi_CHBE','phi_HE','Gmax_LE','Gmax_CLBE','Gmax_BE','Gmax_CHBE','Gmax_HE','M0','OCR','K0','k','status','responsible','inserted_by','timestamp'])
print('==================================================================')
print('Dataframe extracted from mySQL:')
print(df2)
conn.close()

# =============================================================================
#Read of Excel database - PISA sand
# =============================================================================
conn=mysql.connector.connect(user='owdb_user',password='ituotdowdb',host='dklycopilod1',database='owdb')
mycursor=conn.cursor()
mycursor.execute("""SELECT * FROM owdb.COPCAT_Data_Base where instr(name,'sand') and project = 'PISA'""")
db_all = mycursor.fetchall()
df=db_all
    
df3=pd.DataFrame(df, columns=['name','project','location','rev','diameter','length','notes','soil_type','y_u_f','y_u_1','y_u_2','y_u_3','P_u_f','P_u_1','P_u_2','P_u_3','k_p_f','k_p_1','k_p_2','k_p_3','n_p_f','n_p_1','n_p_2','n_p_3','tetam_u_f','tetam_u_1','tetam_u_2','tetam_u_3','m_u_f','m_u_1','m_u_2','m_u_3','k_m_f','k_m_1','k_m_2','k_m_3','n_m_f','n_m_1','n_m_2','n_m_3','yB_u_f','yB_u_1','yB_u_2','yB_u_3','HB_u_f','HB_u_1','HB_u_2','HB_u_3','k_H_f','k_H_1','k_H_2','k_H_3','n_H_f','n_H_1','n_H_2','n_H_3','tetaMb_u_f','tetaMb_u_1','tetaMb_u_2','tetaMb_u_3','MB_u_f','MB_u_1','MB_u_2','MB_u_3','k_MB_f','k_MB_1','k_MB_2','k_MB_3','n_MB_f','n_MB_1','n_MB_2','n_MB_3','preparer','timestamp'])
print('==================================================================')
print('Dataframe extracted from mySQL:')
print(df3)
conn.close()

# =============================================================================
#Read of Excel database - Cowden clay
# =============================================================================
conn=mysql.connector.connect(user='owdb_user',password='ituotdowdb',host='dklycopilod1',database='owdb')
mycursor=conn.cursor()
mycursor.execute("""SELECT * FROM owdb.COPCAT_Data_Base where instr(name,'cowden') and project = 'PISA'""")
db_all = mycursor.fetchall()
df=db_all
    
df4=pd.DataFrame(df, columns=['name','project','location','rev','diameter','length','notes','soil_type','y_u_f','y_u_1','y_u_2','y_u_3','P_u_f','P_u_1','P_u_2','P_u_3','k_p_f','k_p_1','k_p_2','k_p_3','n_p_f','n_p_1','n_p_2','n_p_3','tetam_u_f','tetam_u_1','tetam_u_2','tetam_u_3','m_u_f','m_u_1','m_u_2','m_u_3','k_m_f','k_m_1','k_m_2','k_m_3','n_m_f','n_m_1','n_m_2','n_m_3','yB_u_f','yB_u_1','yB_u_2','yB_u_3','HB_u_f','HB_u_1','HB_u_2','HB_u_3','k_H_f','k_H_1','k_H_2','k_H_3','n_H_f','n_H_1','n_H_2','n_H_3','tetaMb_u_f','tetaMb_u_1','tetaMb_u_2','tetaMb_u_3','MB_u_f','MB_u_1','MB_u_2','MB_u_3','k_MB_f','k_MB_1','k_MB_2','k_MB_3','n_MB_f','n_MB_1','n_MB_2','n_MB_3','preparer','timestamp'])
print('==================================================================')
print('Dataframe extracted from mySQL:')
print(df4)
conn.close()

# =============================================================================
#Choose a position 
# =============================================================================
position='WTG-08'
diameter='9.0'
length='33.0'
write_excel='No' #'Yes' or 'No'

# =============================================================================
#Extract design profile for the position
# =============================================================================
array_df1=df1.to_numpy()
array_df2=df2.to_numpy()
array_df3=df3.to_numpy()
array_df4=df4.to_numpy()

k=0
for index_df2 in range(len(df2)):
    if df2['id'][index_df2]!=position:
        array_df2=np.delete(array_df2,k,0)
    else:
        k=k+1  

# =============================================================================
#Assign PISA parameters
# =============================================================================
array_PISA=[]
index_PISA=0
for index_PISA in range(0,len(array_df2),2):
    array_df3=df3.to_numpy()
    array_df4=df4.to_numpy()
    array_to_supress_sand=array_df3
    array_to_supress_clay=array_df4
    
    if (math.isnan(array_df2[index_PISA][26])):
        dr_position='NaN'
    else:  
        dr_position=int(array_df2[index_PISA][26])
        
    str_dr=str(dr_position)
    str_PISA='Sand_'+str_dr
    
    index_PISA_2=0
    for index_PISA_2 in range(len(array_to_supress_sand)):
        if array_to_supress_sand[index_PISA_2][0] == str_PISA:
            array_PISA=array_PISA+[array_to_supress_sand[index_PISA_2]]
            array_PISA[int(index_PISA/2)][0]=str(array_PISA[int(index_PISA/2)][0]+'_'+str(int(index_PISA/2)))
    
    if str_PISA == 'Sand_NaN':
        array_PISA=array_PISA+[array_to_supress_clay[0]]
        array_PISA[int(index_PISA/2)][0]=str(array_PISA[int(index_PISA/2)][0]+'_'+str(int(index_PISA/2)))
        
array_PISA_non_modified=array_PISA

# =============================================================================
#Assign PU from regression
# =============================================================================
new_array=[]
for index_param in range(0,len(array_df2),2):
    #print('index_param: '+str(index_param))
    for index_df1 in range(len(array_df1)):
        if array_df1[index_df1][1] == array_df2[index_param][4]:
            #print('index_df1: '+str(index_df1))
            #print(array_df1[index_df1])
            
            if(array_df1[index_df1][6]=='z_D'):
                g_dep=array_df2[index_param][3]/float(diameter)
            elif(array_df1[index_df1][6]=='z_L'):
                g_dep=array_df2[index_param][3]/float(length)
            elif(array_df1[index_df1][6]=='L_D'):
                g_dep=float(length)/float(diameter)
            else:
                print('Error found regarding the g_dependant!')
                sys.exit()
                
            if(array_df1[index_df1][4]=='Su'):
                f_dep=array_df2[index_param][22]
            elif(array_df1[index_df1][4]=='tan_phi'):
                f_dep=math.tan(radians(array_df2[index_param][28]))
            elif(array_df1[index_df1][4]=='Dr'):
                f_dep=array_df2[index_param][26]/100
            else:
                print('Error found regarding the f_dependant!')
                sys.exit()
            
            type_f=array_df1[index_df1][3]
            type_g=array_df1[index_df1][5]
            
            if type_f==0 and type_g==0:
                Level_1_1=array_df1[index_df1][8]
                Level_1_2=None
                Level_1_3=None
                Level_2_1=None
                Level_2_2=None
                Level_2_3=None
                Level_3_1=None
                Level_3_2=None
                Level_3_3=None
                
                pug=type_g
                pu1=Level_1_1
                pu2=0
                pu3=0

            elif type_f==0 and type_g==1:
                Level_1_1=array_df1[index_df1][8]
                Level_1_2=None
                Level_1_3=None
                Level_2_1=array_df1[index_df1][11]
                Level_2_2=None
                Level_2_3=None
                Level_3_1=None
                Level_3_2=None
                Level_3_3=None

                pug=type_g
                pu1=Level_1_1
                pu2=Level_2_1
                pu3=0
                
            elif type_f==0 and type_g==2:
                Level_1_1=array_df1[index_df1][8]
                Level_1_2=None
                Level_1_3=None
                Level_2_1=array_df1[index_df1][11]
                Level_2_2=None
                Level_2_3=None
                Level_3_1=array_df1[index_df1][14]
                Level_3_2=None
                Level_3_3=None
                
                pug=type_g
                pu1=Level_1_1
                pu2=Level_2_1
                pu3=Level_3_1

            elif type_f==1 and type_g==0:
                Level_1_1=array_df1[index_df1][8]
                Level_1_2=array_df1[index_df1][9]
                Level_1_3=None
                Level_2_1=None
                Level_2_2=None
                Level_2_3=None
                Level_3_1=None
                Level_3_2=None
                Level_3_3=None
                
                pug=type_g
                pu1=float(Level_1_1)+float(Level_1_2)*float(f_dep)
                pu2=0
                pu3=0

            elif type_f==1 and type_g==1:
                Level_1_1=array_df1[index_df1][8]
                Level_1_2=array_df1[index_df1][9]
                Level_1_3=None
                Level_2_1=array_df1[index_df1][11]
                Level_2_2=array_df1[index_df1][12]
                Level_2_3=None
                Level_3_1=None
                Level_3_2=None
                Level_3_3=None
                
                pug=type_g
                pu1=float(Level_1_1)+float(Level_1_2)*float(f_dep)
                pu2=float(Level_2_1)+float(Level_2_2)*float(f_dep)
                pu3=0
                

            elif type_f==1 and type_g==2:
                Level_1_1=array_df1[index_df1][8]
                Level_1_2=array_df1[index_df1][9]
                Level_1_3=None
                Level_2_1=array_df1[index_df1][11]
                Level_2_2=array_df1[index_df1][12]
                Level_2_3=None
                Level_3_1=array_df1[index_df1][14]
                Level_3_2=array_df1[index_df1][15]
                Level_3_3=None
                
                pug=type_g
                pu1=float(Level_1_1)+float(Level_1_2)*float(f_dep)
                pu2=float(Level_2_1)+float(Level_2_2)*float(f_dep)
                pu3=float(Level_3_1)+float(Level_3_2)*float(f_dep)

            elif type_f==2 and type_g==0:
                Level_1_1=array_df1[index_df1][8]
                Level_1_2=array_df1[index_df1][9]
                Level_1_3=array_df1[index_df1][10]
                Level_2_1=None
                Level_2_2=None
                Level_2_3=None
                Level_3_1=None
                Level_3_2=None
                Level_3_3=None
                
                pug=type_g
                pu1=float(Level_1_1)+float(Level_1_2)*np.exp(float(Level_1_3)*float(f_dep))
                pu2=0
                pu3=0

            elif type_f==2 and type_g==1:
                Level_1_1=array_df1[index_df1][8]
                Level_1_2=array_df1[index_df1][9]
                Level_1_3=array_df1[index_df1][10]
                Level_2_1=array_df1[index_df1][11]
                Level_2_2=array_df1[index_df1][12]
                Level_2_3=array_df1[index_df1][13]
                Level_3_1=None
                Level_3_2=None
                Level_3_3=None
                
                pug=type_g
                pu1=float(Level_1_1)+float(Level_1_2)*np.exp(float(Level_1_3)*float(f_dep))
                pu2=float(Level_2_1)+float(Level_2_2)*np.exp(float(Level_2_3)*float(f_dep))
                pu3=0

            elif type_f==2 and type_g==2:
                Level_1_1=array_df1[index_df1][8]
                Level_1_2=array_df1[index_df1][9]
                Level_1_3=array_df1[index_df1][10]
                Level_2_1=array_df1[index_df1][11]
                Level_2_2=array_df1[index_df1][12]
                Level_2_3=array_df1[index_df1][13]
                Level_3_1=array_df1[index_df1][14]
                Level_3_2=array_df1[index_df1][15]
                Level_3_3=array_df1[index_df1][16]
                
                pug=type_g
                pu1=float(Level_1_1)+float(Level_1_2)*np.exp(float(Level_1_3)*float(f_dep))
                pu2=float(Level_2_1)+float(Level_2_2)*np.exp(float(Level_2_3)*float(f_dep))
                pu3=float(Level_3_1)+float(Level_3_2)*np.exp(float(Level_3_3)*float(f_dep))
            
            new_array=new_array+[pug,pu1,pu2,pu3]
            break
     

for index_new in range(int(len(new_array)/4)):
    array_PISA[index_new][12]=new_array[index_new*4]
    array_PISA[index_new][13]=new_array[index_new*4+1]
    array_PISA[index_new][14]=new_array[index_new*4+2]
    array_PISA[index_new][15]=new_array[index_new*4+3]


# =============================================================================
#Create excel
# =============================================================================
if(write_excel=='Yes'):
    wb = openpyxl.load_workbook('Back_calculation.xlsx')
    #ws=wb.get_active_sheet()
    ws1=wb.create_sheet()
    ws1.title=position
    
    
    max_array_r=len(array_PISA)
    max_array_c=74
    
    for index_row in range(max_array_r):
        for index_column in range(max_array_c):
            ws1.cell(row=index_row+1, column=index_column+2).value=array_PISA[index_row][index_column]
            ws1.cell(row=index_row+1, column=1).value=array_df2[index_row*2+1][4]
            
    wb.save('Back_calculation.xlsx')
    wb.close()

# =============================================================================
#Write DB
# =============================================================================
my_file_kid='Geotec'
my_path = os.path.dirname(__file__)
full_path_kid=(my_path+'/'+my_file_kid+'.gif')

layout = [ [sg.Text('Save the parameters to the database for this position?')],
           
           #[sg.Image(full_path)],
           #[sg.Text('Source for Files', size=(15, 1)), sg.InputText(), sg.FilesBrowse()],
           [sg.Text(str('WTG number: '+position))],
           [sg.Text(str('Diameter: '+diameter))],
           [sg.Text(str('Pile length: '+length))],
           [sg.Text(str('Exemple of name for DB: GHS_U3'))],
           [sg.Text('Naming DB: '),sg.Input(k='-IN-'),sg.Text(size=(14,1), key='-OUT-')],
           [sg.Text('Project: '),sg.Input(k='-IN-0'),sg.Text(size=(14,1), key='-OUT-0')],
           [sg.Text('Location_DD: '),sg.Input(k='-IN-1'),sg.Text(size=(14,1), key='-OUT-1')],
           [sg.Text('Revision: '),sg.Input(k='-IN-2'),sg.Text(size=(14,1), key='-OUT-2')],
           [sg.Text('Preparer: '),sg.Input(k='-IN-3'),sg.Text(size=(14,1), key='-OUT-3')],
           [sg.Button('Save Table in DB')],
           [sg.Image(full_path_kid, key='_IMAGE_')]]

layout = layout 
window = sg.Window('Final prediction',layout,finalize=True)

while True: 
    event, values = window.read(timeout=25)
    if event in (None, 'Cancel'):
        break
    window.Element('_IMAGE_').UpdateAnimation(full_path_kid, time_between_frames=50)
    
    if event == sg.WIN_CLOSED :
        save_table='No'
        break
    
    if event == 'Save Table in DB' :
        save_table='Yes'
        break
    
    window['-OUT-'].update(values['-IN-'])

window.close()

if save_table=='Yes':
    print('Results saved.')
    Name_DB=values['-IN-']
    Project_name=values['-IN-0']
    Location_name=values['-IN-1']
    Revision_number=values['-IN-2']
    Preparer=values['-IN-3']
    Date=datetime.now().date()
    
    conn=mysql.connector.connect(user='owdb_user',password='ituotdowdb',host='dklycopilod1',database='owdb')
    mycursor=conn.cursor()
    
    for index_DB in range(len(array_PISA)):
        sql=str('INSERT INTO COPCAT_Data_Base '+'(name,project,location,rev,diameter,length,notes,soil_type,y_u_F,y_u_1,y_u_2,y_u_3,P_u_F,P_u_1,P_u_2,P_u_3,k_p_F,k_p_1,k_p_2,k_p_3,n_p_F,n_p_1,n_p_2,n_p_3,tetam_u_F,tetam_u_1,tetam_u_2,tetam_u_3,m_u_F,m_u_1,m_u_2,m_u_3,k_m_F,k_m_1,k_m_2,k_m_3,n_m_F,n_m_1,n_m_2,n_m_3,yB_u_F,yB_u_1,yB_u_2,yB_u_3,HB_u_F,HB_u_1,HB_u_2,HB_u_3,k_H_F,k_H_1,k_H_2,k_H_3,n_H_F,n_H_1,n_H_2,n_H_3,tetaMb_u_F,tetaMb_u_1,tetaMb_u_2,tetaMb_u_3,MB_u_F,MB_u_1,MB_u_2,MB_u_3,k_MB_F,k_MB_1,k_MB_2,k_MB_3,n_MB_F,n_MB_1,n_MB_2,n_MB_3,preparer,timestamp) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)')
        #val = str(sql + "INSERT INTO customers (name, address) VALUES (%s, %s)")
        val = (str(str(Name_DB)+'_'+str(index_DB+1)), str(Project_name), str(Location_name), int(Revision_number), float(diameter), str(length), str(array_PISA[index_DB][6]), str(array_PISA[index_DB][7]), array_PISA[index_DB][8], array_PISA[index_DB][9], array_PISA[index_DB][10], array_PISA[index_DB][11], array_PISA[index_DB][12], array_PISA[index_DB][13], array_PISA[index_DB][14], array_PISA[index_DB][15], array_PISA[index_DB][16], array_PISA[index_DB][17], array_PISA[index_DB][18], array_PISA[index_DB][19], array_PISA[index_DB][20], array_PISA[index_DB][21], array_PISA[index_DB][22], array_PISA[index_DB][23], array_PISA[index_DB][24], array_PISA[index_DB][25], array_PISA[index_DB][26], array_PISA[index_DB][27], array_PISA[index_DB][28], array_PISA[index_DB][29], array_PISA[index_DB][30], array_PISA[index_DB][31], array_PISA[index_DB][32], array_PISA[index_DB][33], array_PISA[index_DB][34], array_PISA[index_DB][35], array_PISA[index_DB][36], array_PISA[index_DB][37], array_PISA[index_DB][38], array_PISA[index_DB][39], array_PISA[index_DB][40], array_PISA[index_DB][41], array_PISA[index_DB][42], array_PISA[index_DB][43], array_PISA[index_DB][44], array_PISA[index_DB][45], array_PISA[index_DB][46], array_PISA[index_DB][47], array_PISA[index_DB][48], array_PISA[index_DB][49], array_PISA[index_DB][50], array_PISA[index_DB][51], array_PISA[index_DB][52], array_PISA[index_DB][53], array_PISA[index_DB][54], array_PISA[index_DB][55], array_PISA[index_DB][56], array_PISA[index_DB][57], array_PISA[index_DB][58], array_PISA[index_DB][59], array_PISA[index_DB][60], array_PISA[index_DB][61], array_PISA[index_DB][62], array_PISA[index_DB][63], array_PISA[index_DB][64], array_PISA[index_DB][65], array_PISA[index_DB][66], array_PISA[index_DB][67], array_PISA[index_DB][68], array_PISA[index_DB][69], array_PISA[index_DB][70], array_PISA[index_DB][71], Preparer, Date)
        
        mycursor.execute(sql, val)
        
        conn.commit()
        
        print(mycursor.rowcount, "record inserted.")
    
else:
    print('Results not saved.')
    

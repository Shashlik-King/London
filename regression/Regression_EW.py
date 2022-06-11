# -*- coding: utf-8 -*-
"""
Created on Tue May  3 08:45:27 2022

@author: AYBR
"""
import PySimpleGUI as sg
import mysql.connector
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import math
import pandas as pd
from sklearn.linear_model import LinearRegression 
from math import *
import os, xlsxwriter
import array
import os
from os import walk
import math
import sys
import ctypes
from datetime import datetime, date, timezone
from collections import Counter
import xlsxwriter
import openpyxl
import shutil

#f function
def f(dependant, type_function, variable):
    
    #constant function
    if type_function==0:
        value=variable
    
    #linear function
    elif type_function==1:
        value=variable[0]+variable[1]*dependant
    
    #exponential function
    elif type_function==2:
        value=variable[0]+variable[1]*np.exp(variable[2]*dependant)

    #error type
    else:
        print('Function f type not correctly defined.')
    
    return value


#g function
def g2(input_data, variable1, variable2, variable3, variable4, variable5, variable6, variable7, variable8, variable9):
    
    #define variables and constants
    type_function_g=input_data['type_function_g']
    type_function_f=input_data['type_function_f']
    dependant_g=input_data['dependant_g']
    dependant_f=input_data['dependant_f']
    
    type_function_g=np.array(type_function_g,dtype='float64')
    type_function_f=np.array(type_function_f,dtype='float64')
    dependant_g=np.array(dependant_g,dtype='float64')
    dependant_f=np.array(dependant_f,dtype='float64')

    variable1=np.array(variable1,dtype='float64')
    variable2=np.array(variable2,dtype='float64')
    variable3=np.array(variable3,dtype='float64')
    variable4=np.array(variable4,dtype='float64')
    variable5=np.array(variable5,dtype='float64')
    variable6=np.array(variable6,dtype='float64')
    variable7=np.array(variable7,dtype='float64')
    variable8=np.array(variable8,dtype='float64')
    variable9=np.array(variable9,dtype='float64')
    
    #define the amount of variable in function for f type function
    if(type_function_f==0):
        variable_f_1=[variable1]
        variable_f_2=[variable2]
        variable_f_3=[variable3]
        
    elif(type_function_f==1):    
        variable_f_1=[variable1,variable2]
        variable_f_2=[variable3,variable4]
        variable_f_3=[variable5,variable6]
        
    else:
        variable_f_1=[variable1,variable2,variable3]
        variable_f_2=[variable4,variable5,variable6]
        variable_f_3=[variable8,variable7,variable9]
        
        
    dependant_f=np.array(dependant_f,dtype='float64')
    dependant_g=np.array(dependant_g,dtype='float64')
    
    #e.g.: Pu_1        
    Constant_1=f(dependant_f, type_function_f, variable_f_1)

    #e.g.: Pu_2
    if (type_function_g!=0):
        Constant_2=f(dependant_f, type_function_f, variable_f_2)
        
    #e.g.: Pu_3
    if (type_function_g!=0 and type_function_g!=1):
        Constant_3=f(dependant_f, type_function_f, variable_f_3)
        
    #constant function
    if type_function_g==0:
        value=Constant_1
    
    #linear function
    elif type_function_g==1:
        value=Constant_1+Constant_2*dependant_g
    
    #exponential function
    elif type_function_g==2:
        value=Constant_1+Constant_2*np.exp(Constant_3*dependant_g)
      
    #error type
    else:
        print('Function g type not correctly defined.')
    
    return value


#error function
def Error_function(param_regression,input_data,poptM1,poptM2,poptM3,poptM4,poptM5,poptM6,poptM7,poptM8,poptM9):
    
    max_param_regression=np.amax(param_regression)
    error = (param_regression - g2(input_data,poptM1,poptM2,poptM3,poptM4,poptM5,poptM6,poptM7,poptM8,poptM9))**2/(max_param_regression**2)
    sum_error=np.sum(error)
    return sum_error


#Main function
class Main:  

    print('============================Start=================================')
    plt.close('all')

def __init__(self): 
        self
# =============================================================================
#Read of MYSQL database
# =============================================================================
conn=mysql.connector.connect(user='owdb_user',password='ituotdowdb',host='dklycopilod1',database='owdb')
mycursor=conn.cursor()
mycursor.execute("""SELECT * FROM owdb.Pisa_Param""")
db_all = mycursor.fetchall()
df=db_all
    
df1=pd.DataFrame(df, columns=['Unit_name','Cons_Name','Project_name','Location_name','Rev','Layer_number','Depth','Pile_length','Pile_Diamter','Yu','Pu','Kp','np','Teta','Mu','Km','nm','Yb','Hb','Kb','nb','TetaMb','Mb','kMb','nMb','ASR','Dr','phi','Su','G0','SigmaV','CSR','Neq','CF_OCR','CF_Dr','Contour_diagram_name'])
print('==================================================================')
print('Dataframe extracted from mySQL:')
print(df1)
conn.close()
    
# =============================================================================
#GUI construction
# =============================================================================
layout = [ [sg.Text('Input table for regression')],
           [sg.Text('Name'),sg.Input(k='-IN-'),sg.Text(size=(14,1), key='-OUT-')],
           [sg.Text('Date  '),sg.Input(k='-IN-'),sg.Text(size=(14,1), key='-OUT-')],
           [sg.Text('Source of the File '), sg.FilesBrowse()],
           #[sg.Text('Source for Files', size=(15, 1)), sg.InputText(), sg.FilesBrowse()],
           [sg.Button('Save Table'), sg.Button('Import Table'), sg.Button('Ready')] ]


headings = ['Unit', 'Parameter', 'Automatic','Type g()','Type f()', 'Dependant', 'Model','Batch cyclic']
header =  [[sg.Text('  ')] + [sg.Text(h, size=(19,1)) for h in headings]]

#Lists of parameters for each column
list1=['SS1','H1','H2','H3','PC1','PC2','PH1','PH2','PM1','PM2','PM3','PM4','PM5','CC1','CC2','CC3']
list2=['Yu','Pu','Kp','np','Teta','Mu','Km','nm','Yb','Hb','Kb','nb','TetaMb','Mb','kMb','nMb']
list3=['0','1']
list4=['0','1','2']
list5=['0','1','2']
list6=['Dr','tan_phi','Su','CSR','CF']
list7=['ALL','NGI','GHS']
list8=['ALL','A_FC1.4_OCR1_DR70','A_FC1.4_OCR1_DR83','A_FC1.4_OCR1_DR83_undrained','B_FC15.1_OCR1_DR69_undrained','Batch3_FC14_OCR1_DR77','Batch3_FC14_OCR6_DR80','19b_FC30_OCR6_DR50','20a_FC18_OCR6_DR80','21a_FC11_OCR1.3_DR35','21a_FC17_OCR1.5_DR85','Unit2a_OCR16_Ip16','Unit2c_OCR8_Ip37','C2_OCR6_Ip55','Drammen_Clay_4_PC2','Drammen_Clay_4_PH2','Drammen_Clay_4_PM2','Drammen_Clay_4_CC3']

input_rows = [[sg.Combo(list1, size=(20,1)),sg.Combo(list2, size=(20,1)),sg.Combo(list3, size=(20,1)),sg.Combo(list4, size=(20,1)),sg.Combo(list5, size=(20,1)),sg.Combo(list6, size=(20,1)),sg.Combo(list7, size=(20,1)),sg.Combo(list8, size=(20,1))] for row in range(10)]
layout = layout + header + input_rows 
window = sg.Window('Regression',layout,finalize=True)


#Read GUI values
while True: 
    event, values = window.read()
    
    if event:
        window['-OUT-'].update(values)
        if values['-IN-0']=='':
            values['-IN-0']=datetime.now().date()
            window['-IN-0'].Update(values['-IN-0'])
        if values['-IN-']=='':
            values['-IN-']='Regression'
            window['-IN-'].Update(values['-IN-'])

    if event == sg.WIN_CLOSED or event == 'Ready':
        break
    window['-OUT-'].update(values['-IN-'])
    
    if event == 'Save Table':
        print('Saved Table!')
        title_txt=(str(values['-IN-'])+'_'+str(values['-IN-0'])+'.txt')
        txt_file=open(str(title_txt),'w+')
        
        #txt_file.writelines('|')
        txt_file.writelines(str(values['-IN-']))
        #txt_file.writelines('\n')
        txt_file.writelines('|')
        txt_file.writelines(str(values['-IN-0']))
        #txt_file.writelines('\n')
        txt_file.writelines('|')
        txt_file.writelines(str(values['Browse']))
        #txt_file.writelines('\n')
        
        for index_write_txt in range(len(values)-3):
            txt_file.writelines('|')
            txt_file.writelines(str(values[index_write_txt]))
            #txt_file.writelines('\n')

        txt_file.close() 
        
    if event == 'Import Table':
        if values['Browse']=='':
            print('No table selected!')
        else:
            txt_file=open(values['Browse'],'r+')  
            content=txt_file.read()
            print(content)
            counter_str=Counter(content)
            number_count=counter_str['|']
            print(number_count)
            
            content_split=content.split('|')
            print(content_split)
            print(range(len(content_split)))
            
            window['-IN-'].Update(content_split[0])
            window['-IN-0'].Update(content_split[1])
            window['Browse'].Update(content_split[2])
            
            index_write_table=0
            for index_write_table in range(number_count-2):
                window[index_write_table].Update(content_split[index_write_table+3])
            
            txt_file.close() 
           
window.close()

#The GUI can be ignored filling the line below
###############################################################################
#values={'-IN-': '', '-IN-0': '', 0: 'SS1', 1: 'Pu', 2: '0', 3: '0', 4: '2', 5: 'Dr', 6: 'ALL', 7: 'ALL', 8: 'PC1', 9: 'Pu', 10: '0', 11: '1', 12: '1', 13: 'Dr', 14: 'GHS', 15: 'ALL', 16: '', 17: '', 18: '', 19: '', 20: '', 21: '', 22: '', 23: '', 24: '', 25: '', 26: '', 27: '', 28: '', 29: '', 30: '', 31: '', 32: '', 33: '', 34: '', 35: '', 36: '', 37: '', 38: '', 39: '', 40: '', 41: '', 42: '', 43: '', 44: '', 45: '', 46: '', 47: '', 48: '', 49: '', 50: '', 51: '', 52: '', 53: '', 54: '', 55: '', 56: '', 57: '', 58: '', 59: '', 60: '', 61: '', 62: '', 63: '', 64: '', 65: '', 66: '', 67: '', 68: '', 69: '', 70: '', 71: '', 72: '', 73: '', 74: '', 75: '', 76: '', 77: '', 78: '', 79: ''}
###############################################################################

# =============================================================================
#Table verification
# =============================================================================
inputs_regression = values

#check of the 10 lines if units empty or filled
list_regression = [0,8,16,24,32,40,48,56,64,72]
list_toDo = []

for i in range(len(list_regression)):
    if values[list_regression[i]]!='':
        list_toDo=list_toDo+[i]

#Check where we have units for the regression
if list_toDo==[]:
    print('==================================================================')
    print ('No unit selected for the regression!')
    sys.exit()

#Check if we have all column filled for the units used for the regression
for i in range(len(list_toDo)):
    if values[list_toDo[i]+1] == '' or values[list_toDo[i]+2] == '' or values[list_toDo[i]+3] == '' or values[list_toDo[i]+4] == '' or values[list_toDo[i]+5] == '' or values[list_toDo[i]+6] == '' or values[list_toDo[i]+7] == '' :
        print('==================================================================')
        print ('Missing info to run the regression - empty column!')
        sys.exit()

#Set up of naming and date in case of empty filling
if values['-IN-']=='':
    name_regression=('Default_naming')
else :
    name_regression=values['-IN-']
    
if values['-IN-0']=='':
    date_project='01-01-2022'
else :
    date_project=values['-IN-0']

    
# =============================================================================
#Regression
# =============================================================================
#Transform dataframe of mysql in a simple array
array_regression_base=df1.to_numpy()

#Batch run of the regressions for each line fully filled
for i in range(len(list_toDo)):
    array_regression=array_regression_base
    print('i:',i)
    print(array_regression)
    print('==================================================================')
    print('Start of the regression:')
    
    #Lookup at if 'Automatic' is set to 0 or 1 - if 1 run all possibilities
    if values[list_toDo[i]*8+2]=='1' :
        print('==================================================================')
        print('Automatic regression selected!')
        
        k=0
        for j in range(len(df1)):
            #print(j)
            if df1['Unit_name'].values[j]==values[list_toDo[i]*8] and (df1['Cons_Name'].values[j]==values[list_toDo[i]*8+6] or values[list_toDo[i]*8+6]=='ALL') and (df1['Contour_diagram_name'].values[j]==values[list_toDo[i]*8+7] or values[list_toDo[i]*8+7]=='ALL'):
                k=k+1
                Cons_name_excel=values[list_toDo[i]*8+6]
            else:
                #remove the lines that do not validate: unit / model / contour chosen
                array_regression=np.delete(array_regression,k,0)
        
        Unit=array_regression[0]
        
        #Batch run over the Sand parameters and all types of function f and g
        if(Unit[0]=='SS1' or Unit[0]=='H2' or Unit[0]=='PC1' or Unit[0]=='PH1' or Unit[0]=='PM1' or Unit[0]=='PM3' or Unit[0]=='PM4' or Unit[0]=='PM5' or Unit[0]=='CC1'):
            print('==================================================================')
            print('Running on parameters of SAND: Dr, tan(phi), CSR, G0, Cf')
            list_sand=['Dr','tan_phi','CSR','CF']
            list_f=[0,1,2]
            list_g=[0,1,2]
            index_sand=0
            index_g=0
            index_f=0
            index_table=0
            index_excel_count=0
            error_table=[]
            error_max=[]
            error_final_to_show=[]
            for index_sand in range(len(list_sand)):
                for index_g in range(len(list_g)):
                    for index_f in range(len(list_f)):
                        if((index_f>=3 and index_g>=3)):
                            print('Error!')
                        else:
                            # parameters set up
                            dependant_f_str=list_sand[index_sand]
                            type_f=list_f[index_f]
                            type_g=list_g[index_g]
                            param_regression_str=values[list_toDo[i]*8+1]
                            
                            Unit=array_regression[:,0]
                            Model=array_regression[:,1]
                            Project=array_regression[:,2]
                            Location=array_regression[:,3]
                            Rev=array_regression[:,4]
                            Layer=array_regression[:,5]
                            depth=-array_regression[:,6]
                            length=array_regression[:,7]
                            diameter=array_regression[:,8]
                            
                            YU=array_regression[:,9]
                            PU=array_regression[:,10]
                            KP=array_regression[:,11]
                            NP=array_regression[:,12]
                            TETA=array_regression[:,13]
                            MU=array_regression[:,14]
                            KM=array_regression[:,15]
                            NM=array_regression[:,16]
                            YB=array_regression[:,17]
                            HB=array_regression[:,18]
                            KB=array_regression[:,19]
                            NB=array_regression[:,20]
                            TETAMB=array_regression[:,21]
                            MB=array_regression[:,22]
                            KMB=array_regression[:,23]
                            NMB=array_regression[:,24]
                            
                            ASR=array_regression[:,25]
                            Dr=array_regression[:,26]
                            phi=array_regression[:,27]
                            phi_rad=[0]*(len(phi))
                            for i_phi in range(len(phi)):
                                phi_rad[i_phi]=np.radians(phi[i_phi])
                            Su=array_regression[:,28]  
                            G0=array_regression[:,29]   
                            SigmaV=array_regression[:,30]   
                            CSR=array_regression[:,31]   
                            Neq=array_regression[:,32]   
                            CF_OCR=array_regression[:,33] 
                            CF_Dr=array_regression[:,34] 
                            Batch=array_regression[:,35] 
                            
                            z_L=depth/length
                            z_D=depth/diameter
                            L_D=length/diameter
                            Dr=Dr*0.01
                            tan_phi=np.tan(phi_rad)
                            CF=CF_OCR*CF_Dr
                                    
                            input_data={}
                            input_data['Unit']=Unit
                            input_data['Model']=Model
                            input_data['Project']=Project
                            input_data['Location']=Location
                            input_data['Rev']=Rev
                            input_data['Layer']=Layer
                            input_data['depth']=depth
                            input_data['length']=length
                            input_data['diameter']=diameter
                            
                            input_data['Yu']=YU
                            input_data['Pu']=PU
                            input_data['Kp']=KP
                            input_data['np']=NP
                            input_data['Teta']=TETA
                            input_data['Mu']=MU
                            input_data['Km']=KM
                            input_data['nm']=NM
                            input_data['Yb']=YB
                            input_data['Hb']=HB
                            input_data['Kb']=KB
                            input_data['nb']=NB
                            input_data['TetaMb']=TETAMB
                            input_data['Mb']=MB
                            input_data['kMb']=KMB
                            input_data['nMb']=NMB
                            
                            input_data['ASR']=ASR
                            input_data['Dr']=Dr
                            input_data['phi']=phi
                            input_data['phi_rad']=phi_rad
                            input_data['Su']=Su
                            input_data['G0']=G0
                            input_data['SigmaV']=SigmaV
                            input_data['CSR']=CSR
                            input_data['Neq']=Neq
                            input_data['CF_OCR']=CF_OCR
                            input_data['CF_Dr']=CF_Dr
                            input_data['Batch']=Batch
                            
                            input_data['CF']= CF
                            input_data['z_L']=z_L
                            input_data['z_D']=z_D
                            input_data['L_D']=L_D
                            input_data['tan_phi']=tan_phi
                            
                            input_data['type_function_f']=type_f
                            input_data['type_function_g']=type_g
                            
                            print(Unit[0])
                            print(param_regression_str)
                            
                            #define g dependant as per PISA paper for SAND
                            if(Unit[0]=='SS1' or Unit[0]=='H2' or Unit[0]=='PC1' or Unit[0]=='PH1' or Unit[0]=='PM1' or Unit[0]=='PM3' or Unit[0]=='PM4' or Unit[0]=='PM5' or Unit[0]=='CC1'):
                                if(param_regression_str=='Yb' or param_regression_str=='Hb' or param_regression_str=='Kb' or param_regression_str=='nb'):
                                    input_data['dependant_g']=L_D
                                    dependant_g_str='L_D'
                                else:
                                    input_data['dependant_g']=z_L
                                    dependant_g_str='z_L'
                            else:
                                if(param_regression_str=='Yu' or param_regression_str=='Pu' or param_regression_str=='Kp' or param_regression_str=='np'):
                                    input_data['dependant_g']=z_D
                                    dependant_g_str='z_D'
                                elif (param_regression_str=='Teta' or param_regression_str=='Mu' or param_regression_str=='Km' or param_regression_str=='nm'):
                                    input_data['dependant_g']=z_L
                                    dependant_g_str='z_L'
                                else:
                                    input_data['dependant_g']=L_D
                                    dependant_g_str='L_D'
                                
                            input_data['dependant_f']=input_data[dependant_f_str]
                            param_regression=input_data[param_regression_str]
                            
                            #print('==================================================================')
                            #print(range(len(PISA)))
                            #print(type(PISA))
                            try: 
                                poptM, pcovM = curve_fit(g2,input_data, param_regression)
                                #print(poptM)
                                #print(variable_total)
                                predict_param=g2(input_data,poptM[0],poptM[1],poptM[2],poptM[3],poptM[4],poptM[5],poptM[6],poptM[7],poptM[8]) 
                                
                                index_constant=0
                                if(type_g==0 and type_f==0):
                                    constant=predict_param
                                    predict_param=[]
                                    for index_constant in range(len(param_regression)):
                                        predict_param=predict_param+[constant]
                               
                                #print('Prediction: ', predict_param)
                                
                                depth_profile=np.amax(depth)
                                calc_y=round(depth_profile/5)+1
                                      
                                M_profile=np.amax(param_regression)
                                calc_yM=round(M_profile/100)+1
                                      
                                fig = plt.figure(figsize=(10, 10))
                                ax1 = fig.add_subplot(111)
                                      #ax1.grid()
                                      
                                line1=plt.scatter(param_regression,depth)
                                line1.set_label("Real values")
                                      
                                line6=plt.scatter(predict_param,depth,color='black')
                                line6.set_label('Predicted values')
                                  
                                ax1.set_xlabel(param_regression_str)
                                      # ax1.set_xlim(0,max(Gmax))
                                min_depth=0
                                max_depth=calc_y*5
                                plt.ylim([min_depth,max_depth])
                                plt.xlim([0,calc_yM*M_profile*2])
                                plt.gca().invert_yaxis()
                                ax1.set_ylabel('Depth [m]')
                                      
                                ax1.legend()
                                
                                type_f=str(list_f[index_f])
                                type_g=str(list_g[index_g])
                                
                                if(type_g=='0'):
                                    if(type_f=='0'):
                                        plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+") )")
                                        title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                    elif(type_f=='1'):
                                        plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+ "*"+dependant_f_str +") )")
                                        title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                    else:
                                        plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+"*exp("+str(round(poptM[2],2))+"*"+dependant_f_str +") )")
                                        title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                elif(type_g=='1'):
                                    if(type_f=='0'):
                                        plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[1],2))+"))"+" *"+dependant_g_str+" )")
                                        title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                    elif(type_f=='1'):
                                        plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+ "*"+dependant_f_str +") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[2],2))+"+"+str(round(poptM[3],2))+ "*"+dependant_f_str+"))"+" *"+dependant_g_str+" )")
                                        title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                    else:
                                        plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+"*exp("+str(round(poptM[2],2))+ "*"+dependant_f_str +") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[3],2))+"+"+str(round(poptM[4],2))+"*exp("+str(round(poptM[5],2))+ "*"+dependant_f_str+"))"+" *"+dependant_g_str+" )")
                                        title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                else:
                                     if(type_f=='0'):
                                         plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[1],2))+"))"+"*exp("+param_regression_str+"_3= ("+str(round(poptM[2],2))+"))"+" *"+dependant_g_str+") )")
                                         title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                     elif(type_f=='1'):
                                         plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+ "*"+dependant_f_str +") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[2],2))+"+"+str(round(poptM[3],2))+ "*"+dependant_f_str+"))"+"*exp("+param_regression_str+"_3= ("+str(round(poptM[4],2))+"+"+str(round(poptM[5],2))+"*"+dependant_f_str+")"+" *"+dependant_g_str+") )")
                                         title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                     else:
                                         plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+"*exp("+str(round(poptM[2],2))+ "*"+dependant_f_str +") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[3],2))+"+"+str(round(poptM[4],2))+"*exp("+str(round(poptM[5],2))+ "*"+dependant_f_str+"))"+"*exp("+param_regression_str+"_3= ("+str(round(poptM[6],2))+"+"+str(round(poptM[7],2))+"*exp("+str(round(poptM[8],2))+"*"+dependant_f_str+"))"" *"+dependant_g_str+") )")
                                         title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                            
                                      
                                #variable_total= [0.3667, 25.89, 0.3375, -8.9, 1, 1]     
                                my_path = os.path.dirname(__file__)
                                my_file = title_plot
                                subdirectory=('\Plots_'+values['-IN-']+'_'+values['-IN-0'])
                                full_path=(my_path+subdirectory+'/'+my_file+'.png')
                                print(my_path)
                                print(my_file)
                                print(subdirectory)
                                if not os.path.isdir(my_path+subdirectory):
                                    os.makedirs(my_path+subdirectory)
                        
                                fig.savefig(full_path)
                                
                                title_excel_original='\Template.xlsx'
                                title_excel_target='\Regression_parameters.xlsx'
                                
                                original=(my_path+title_excel_original)
                                target=(my_path+subdirectory+title_excel_target)
                                
                                if not os.path.isfile(my_path+subdirectory+title_excel_target):
                                    shutil.copyfile(original, target)
                                
                                my_file_excel=str(Unit[0]+"_"+param_regression_str+"_"+Cons_name_excel)
                                
                                workbook = openpyxl.load_workbook(target)
                                if (my_file_excel in workbook.sheetnames) and index_excel_count==0:
                                    workbook.remove_sheet(workbook.get_sheet_by_name(my_file_excel))
                                    workbook.create_sheet(my_file_excel)
                                elif index_excel_count==0: 
                                    workbook.create_sheet(my_file_excel)
                                
                                sheet1=workbook['Template']
                                sheet2=workbook[my_file_excel]
                                
                                if index_excel_count==0:
                                

                                    maxr = sheet1.max_row
                                    maxc = sheet1.max_column
                        
                                    for r in range (1, maxr + 1):
                                        for c in range (1, maxc + 1):
                                            sheet2.cell(row=r,column=c).value = sheet1.cell(row=r,column=c).value
                                    
                                    index_row=0
                                    index_column=0
                                    
                                    max_array_r=len(array_regression)
                                    max_array_c=36
                                    
                                    for index_row in range(max_array_r):
                                        for index_column in range(max_array_c):
                                            sheet2.cell(row=index_row+3, column=index_column+1).value=array_regression[index_row][index_column]
                                            
                                
                                index_variable=0
                                
                                for index_variable in range(len(poptM)):
                                    sheet2.cell(row=index_excel_count+3, column=index_variable+44).value=poptM[index_variable]
                                
                                sheet2.cell(row=index_excel_count+3, column=54).value=dependant_f_str
                                sheet2.cell(row=index_excel_count+3, column=55).value=dependant_g_str
                                sheet2.cell(row=index_excel_count+3, column=57).value=type_f
                                sheet2.cell(row=index_excel_count+3, column=58).value=type_g
                                
                                sheet2.cell(row=index_excel_count+3, column=38).value=param_regression_str
                                

                                
                                
                                print(index_excel_count)
                                
                                error=Error_function(param_regression,input_data,poptM[0],poptM[1],poptM[2],poptM[3],poptM[4],poptM[5],poptM[6],poptM[7],poptM[8])
                                
                                error_table=error_table+[('Error_',param_regression_str,'_f:',type_f,'_g:',type_g,'_dependant_f:',dependant_f_str,'_dependant_g:',dependant_g_str,'_error:',error)]
                                error_final_to_show=error_final_to_show+[('Variables:',str(poptM[0]),str(poptM[1]),str(poptM[2]),str(poptM[3]),str(poptM[4]),str(poptM[5]),str(poptM[6]),str(poptM[7]),str(poptM[8]))]
                                error_max=error_max+[error]
                                
                                print('==================================================================')
                                print('Error_predict: ', error)  
                                
                                sheet2.cell(row=index_excel_count+3, column=60).value=error
                                workbook.save(target)
                                workbook.close()
                                index_excel_count=index_excel_count+1
                                
                                
                            except:
                                pass
        else:
            print('==================================================================')
            print('Running on parameters of CLAY: Su, CSR, G0, Cf')
            list_clay=['Su','CSR','CF']
            list_f=[0,1,2]
            list_g=[0,1,2]
            index_clay=0
            index_g=0
            index_f=0
            index_table=0
            index_excel_count=0
            error_table=[]
            error_max=[]
            error_final_to_show=[]
            for index_sand in range(len(list_clay)):
                for index_g in range(len(list_g)):
                    for index_f in range(len(list_f)):
                        if(index_f>=3 and index_g>=3):
                            print('Error!')
                        else:
                            
                            # parameters set up
                            dependant_f_str=list_clay[index_clay]
                            type_f=list_f[index_f]
                            type_g=list_g[index_g]
                            param_regression_str=values[list_toDo[i]*8+1]
                            
                            Unit=array_regression[:,0]
                            Model=array_regression[:,1]
                            Project=array_regression[:,2]
                            Location=array_regression[:,3]
                            Rev=array_regression[:,4]
                            Layer=array_regression[:,5]
                            depth=-array_regression[:,6]
                            length=array_regression[:,7]
                            diameter=array_regression[:,8]
                            
                            YU=array_regression[:,9]
                            PU=array_regression[:,10]
                            KP=array_regression[:,11]
                            NP=array_regression[:,12]
                            TETA=array_regression[:,13]
                            MU=array_regression[:,14]
                            KM=array_regression[:,15]
                            NM=array_regression[:,16]
                            YB=array_regression[:,17]
                            HB=array_regression[:,18]
                            KB=array_regression[:,19]
                            NB=array_regression[:,20]
                            TETAMB=array_regression[:,21]
                            MB=array_regression[:,22]
                            KMB=array_regression[:,23]
                            NMB=array_regression[:,24]
                            
                            ASR=array_regression[:,25]
                            Dr=array_regression[:,26]
                            phi=array_regression[:,27]
                            phi_rad=[0]*(len(phi))
                            for i_phi in range(len(phi)):
                                phi_rad[i_phi]=np.radians(phi[i_phi])
                            Su=array_regression[:,28]  
                            G0=array_regression[:,29]   
                            SigmaV=array_regression[:,30]   
                            CSR=array_regression[:,31]   
                            Neq=array_regression[:,32]   
                            CF_OCR=array_regression[:,33] 
                            CF_Dr=array_regression[:,34] 
                            Batch=array_regression[:,35] 
                            
                            z_L=depth/length
                            z_D=depth/diameter
                            L_D=length/diameter
                            Dr=Dr*0.01
                            tan_phi=np.tan(phi_rad)
                            CF=CF_OCR*CF_Dr
                  
                            input_data={}
                            input_data['Unit']=Unit
                            input_data['Model']=Model
                            input_data['Project']=Project
                            input_data['Location']=Location
                            input_data['Rev']=Rev
                            input_data['Layer']=Layer
                            input_data['depth']=depth
                            input_data['length']=length
                            input_data['diameter']=diameter
                            
                            input_data['Yu']=YU
                            input_data['Pu']=PU
                            input_data['Kp']=KP
                            input_data['np']=NP
                            input_data['Teta']=TETA
                            input_data['Mu']=MU
                            input_data['Km']=KM
                            input_data['nm']=NM
                            input_data['Yb']=YB
                            input_data['Hb']=HB
                            input_data['Kb']=KB
                            input_data['nb']=NB
                            input_data['TetaMb']=TETAMB
                            input_data['Mb']=MB
                            input_data['kMb']=KMB
                            input_data['nMb']=NMB
                            
                            input_data['ASR']=ASR
                            input_data['Dr']=Dr
                            input_data['phi']=phi
                            input_data['phi_rad']=phi_rad
                            input_data['Su']=Su
                            input_data['G0']=G0
                            input_data['SigmaV']=SigmaV
                            input_data['CSR']=CSR
                            input_data['Neq']=Neq
                            input_data['CF_OCR']=CF_OCR
                            input_data['CF_Dr']=CF_Dr
                            input_data['Batch']=Batch
                            
                            input_data['CF']= CF
                            input_data['z_L']=z_L
                            input_data['z_D']=z_D
                            input_data['L_D']=L_D
                            input_data['tan_phi']=tan_phi
                            
                            input_data['type_function_f']=type_f
                            input_data['type_function_g']=type_g
                            
                            #print(Unit[0])
                            #print(param_regression_str)
                            
                            #define g dependant as per PISA paper for CLAY
                            if(Unit[0]=='SS1' or Unit[0]=='H2' or Unit[0]=='PC1' or Unit[0]=='PH1' or Unit[0]=='PM1' or Unit[0]=='PM3' or Unit[0]=='PM4' or Unit[0]=='PM5' or Unit[0]=='CC1'):
                                if(param_regression_str=='Yb' or param_regression_str=='Hb' or param_regression_str=='Kb' or param_regression_str=='nb'):
                                    input_data['dependant_g']=L_D
                                    dependant_g_str='L_D'
                                else:
                                    input_data['dependant_g']=z_L
                                    dependant_g_str='z_L'
                            else:
                                if(param_regression_str=='Yu' or param_regression_str=='Pu' or param_regression_str=='Kp' or param_regression_str=='np'):
                                    input_data['dependant_g']=z_D
                                    dependant_g_str='z_D'
                                elif (param_regression_str=='Teta' or param_regression_str=='Mu' or param_regression_str=='Km' or param_regression_str=='nm'):
                                    input_data['dependant_g']=z_L
                                    dependant_g_str='z_L'
                                else:
                                    input_data['dependant_g']=L_D
                                    dependant_g_str='L_D'
                                
                            #print(dependant_g_str)
                            input_data['dependant_f']=input_data[dependant_f_str]
                            param_regression=input_data[param_regression_str]
                            
                            #print('==================================================================')
                            #print(range(len(PISA)))
                            #print(type(PISA))
                            try:
                                poptM, pcovM = curve_fit(g2,input_data, param_regression)
                                #print(poptM)
                                #print(variable_total)
                                predict_param=g2(input_data,poptM[0],poptM[1],poptM[2],poptM[3],poptM[4],poptM[5],poptM[6],poptM[7],poptM[8]) 
                                
                                index_constant=0
                                if(type_g==0 and type_f==0):
                                    constant=predict_param
                                    predict_param=[]
                                    for index_constant in range(len(param_regression)):
                                        predict_param=predict_param+[constant]
                               
                                print('Prediction: ', predict_param)
                                
                                depth_profile=np.amax(depth)
                                calc_y=round(depth_profile/5)+1
                                      
                                M_profile=np.amax(param_regression)
                                calc_yM=round(M_profile/100)+1
                                      
                                fig = plt.figure(figsize=(10, 10))
                                ax1 = fig.add_subplot(111)
                                      #ax1.grid()
                                      
                                line1=plt.scatter(param_regression,depth)
                                line1.set_label("Real values")
                                      
                                line6=plt.scatter(predict_param,depth,color='black')
                                line6.set_label('Predicted values')
                                  
                                ax1.set_xlabel(param_regression_str)
                                      # ax1.set_xlim(0,max(Gmax))
                                min_depth=0
                                max_depth=calc_y*5
                                plt.ylim([min_depth,max_depth])
                                plt.xlim([0,calc_yM*M_profile*2])
                                plt.gca().invert_yaxis()
                                ax1.set_ylabel('Depth [m]')
                                      
                                ax1.legend()
                                
                                type_f=str(list_f[index_f])
                                type_g=str(list_g[index_g])
                                
                                if(type_g=='0'):
                                    if(type_f=='0'):
                                        plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+") )")
                                        title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                    elif(type_f=='1'):
                                        plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+ "*"+dependant_f_str +") )")
                                        title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                    else:
                                        plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+"*exp("+str(round(poptM[2],2))+"*"+dependant_f_str +") )")
                                        title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                elif(type_g=='1'):
                                    if(type_f=='0'):
                                        plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[1],2))+"))"+" *"+dependant_g_str+" )")
                                        title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                    elif(type_f=='1'):
                                        plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+ "*"+dependant_f_str +") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[2],2))+"+"+str(round(poptM[3],2))+ "*"+dependant_f_str+"))"+" *"+dependant_g_str+" )")
                                        title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                    else:
                                        plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+"*exp("+str(round(poptM[2],2))+ "*"+dependant_f_str +") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[3],2))+"+"+str(round(poptM[4],2))+"*exp("+str(round(poptM[5],2))+ "*"+dependant_f_str+"))"+" *"+dependant_g_str+" )")
                                        title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                else:
                                     if(type_f=='0'):
                                         plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[1],2))+"))"+"*exp("+param_regression_str+"_3= ("+str(round(poptM[2],2))+"))"+" *"+dependant_g_str+") )")
                                         title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                     elif(type_f=='1'):
                                         plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+ "*"+dependant_f_str +") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[2],2))+"+"+str(round(poptM[3],2))+ "*"+dependant_f_str+"))"+"*exp("+param_regression_str+"_3= ("+str(round(poptM[4],2))+"+"+str(round(poptM[5],2))+"*"+dependant_f_str+")"+" *"+dependant_g_str+") )")
                                         title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                     else:
                                         plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+"*exp("+str(round(poptM[2],2))+ "*"+dependant_f_str +") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[3],2))+"+"+str(round(poptM[4],2))+"*exp("+str(round(poptM[5],2))+ "*"+dependant_f_str+"))"+"*exp("+param_regression_str+"_3= ("+str(round(poptM[6],2))+"+"+str(round(poptM[7],2))+"*exp("+str(round(poptM[8],2))+"*"+dependant_f_str+"))"" *"+dependant_g_str+") )")
                                         title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                                      
                                #variable_total= [0.3667, 25.89, 0.3375, -8.9, 1, 1]     
                                my_path = os.path.dirname(__file__)
                                my_file = title_plot
                                subdirectory=('\Plots_'+values['-IN-']+'_'+values['-IN-0'])
                                full_path=(my_path+subdirectory+'/'+my_file+'.png')
                                print(my_path)
                                print(my_file)
                                print(subdirectory)
                                if not os.path.isdir(my_path+subdirectory):
                                    os.makedirs(my_path+subdirectory)
                        
                                fig.savefig(full_path)
                                
                                title_excel_original='\Template.xlsx'
                                title_excel_target='\Regression_parameters.xlsx'
                                
                                original=(my_path+title_excel_original)
                                target=(my_path+subdirectory+title_excel_target)
                                
                                if not os.path.isfile(my_path+subdirectory+title_excel_target):
                                    shutil.copyfile(original, target)
                                
                                my_file_excel=str(Unit[0]+"_"+param_regression_str+"_"+Cons_name_excel)
                                
                                workbook = openpyxl.load_workbook(target)
                                if (my_file_excel in workbook.sheetnames) and index_excel_count==0:
                                    workbook.remove_sheet(workbook.get_sheet_by_name(my_file_excel))
                                    workbook.create_sheet(my_file_excel)
                                elif index_excel_count==0: 
                                    workbook.create_sheet(my_file_excel)
                                
                                sheet1=workbook['Template']
                                sheet2=workbook[my_file_excel]
                                
                                if index_excel_count==0:
                                

                                    maxr = sheet1.max_row
                                    maxc = sheet1.max_column
                        
                                    for r in range (1, maxr + 1):
                                        for c in range (1, maxc + 1):
                                            sheet2.cell(row=r,column=c).value = sheet1.cell(row=r,column=c).value
                                    
                                    index_row=0
                                    index_column=0
                                    
                                    max_array_r=len(array_regression)
                                    max_array_c=36
                                    
                                    for index_row in range(max_array_r):
                                        for index_column in range(max_array_c):
                                            sheet2.cell(row=index_row+3, column=index_column+1).value=array_regression[index_row][index_column]
                                            
                                
                                index_variable=0
                                
                                for index_variable in range(len(poptM)):
                                    sheet2.cell(row=index_excel_count+3, column=index_variable+44).value=poptM[index_variable]
                                
                                sheet2.cell(row=index_excel_count+3, column=54).value=dependant_f_str
                                sheet2.cell(row=index_excel_count+3, column=55).value=dependant_g_str
                                sheet2.cell(row=index_excel_count+3, column=57).value=type_f
                                sheet2.cell(row=index_excel_count+3, column=58).value=type_g
                                
                                sheet2.cell(row=index_excel_count+3, column=38).value=param_regression_str
                                
                                                                
                                
                                print(index_excel_count)
                                
                                error=Error_function(param_regression,input_data,poptM[0],poptM[1],poptM[2],poptM[3],poptM[4],poptM[5],poptM[6],poptM[7],poptM[8])
                                
                                error_table=error_table+[('Error_',param_regression_str,'_f:',type_f,'_g:',type_g,'_dependant_f:',dependant_f_str,'_dependant_g:',dependant_g_str,'_error:',error)]
                                error_final_to_show=error_final_to_show+[('Variables:',str(poptM[0]),str(poptM[1]),str(poptM[2]),str(poptM[3]),str(poptM[4]),str(poptM[5]),str(poptM[6]),str(poptM[7]),str(poptM[8]))]
                                error_max=error_max+[error]
                                
                                print('==================================================================')
                                print('Error_predict: ', error)  
                                
                                sheet2.cell(row=index_excel_count+3, column=60).value=error
                                workbook.save(target)
                                workbook.close()
                                index_excel_count=index_excel_count+1
                            except:
                                pass
        print('==================================================================')
        print('Final_error_table:')

        index_table=0
        error_brut=[]
        for index_table in range(len(error_table)):
            print(error_table[index_table])
            error_brut=error_brut+[error_table[index_table][11]];
        
        #for index_brut in range(len(error_brut)):
            
        array_brut=np.array(error_brut)
        order=array_brut.argsort()
        ranks=order.argsort()
        
        print('==================================================================')
        print('Min_error_parameters_associated:')
        max_error=np.min(error_max)
        print(max_error)
        index_error=np.argmin(error_max)
        print(error_table[index_error])
        
        #print(error_table)
        #sys.exit()
        
        for index_check_brut in range(len(ranks)):
            position_brut=np.where(ranks==index_check_brut)[0][0]
            array_full_str=error_table[position_brut]
            array_to_show=error_final_to_show[position_brut]
            
            #Parameters level 1
            Level_1=[float(array_to_show[1]),float(array_to_show[2]),float(array_to_show[3]),float(array_to_show[4]),float(array_to_show[5]),float(array_to_show[6]),float(array_to_show[7]),float(array_to_show[8]),float(array_to_show[9])]
            #dependant_g=input_data['dependant_g']
            #dependant_f=input_data['dependant_f']
            
            #SS1
            check_Dr=[0.52,1]
            check_tan_phi=[0.55,0.93]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[0,1200]
            check_z_L=[0,0.2]
            check_z_D=[0,1]
            check_L_D=[3,5]
            
            #H1
            check_Dr=[0.52,1]
            check_tan_phi=[0.55,0.93]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[0,488]
            check_z_L=[0,0.2]
            check_z_D=[0,1]
            check_L_D=[3,5]
            
            #H2
            check_Dr=[0.22,1]
            check_tan_phi=[0.83,1.07]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[0,1200]
            check_z_L=[0,0.3]
            check_z_D=[0,1.5]
            check_L_D=[3,5]
            
            #H3
            check_Dr=[0.22,1]
            check_tan_phi=[0.83,1.07]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[21,402]
            check_z_L=[0,0.3]
            check_z_D=[0,1.5]
            check_L_D=[3,5]
            
            #PC1
            check_Dr=[0.26,1]
            check_tan_phi=[0.55,0.93]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[21,402]
            check_z_L=[0,1]
            check_z_D=[0,5]
            check_L_D=[3,5]
            
            #PC2
            check_Dr=[0.26,1]
            check_tan_phi=[0.55,0.93]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[24,713]
            check_z_L=[0,1]
            check_z_D=[0,3]
            check_L_D=[3,5]
            
            #PH1
            check_Dr=[0.57,0.84]
            check_tan_phi=[0.55,0.60]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[24,713]
            check_z_L=[0.1,1]
            check_z_D=[1,5]
            check_L_D=[3,5]
            
            #PH2
            check_Dr=[0.57,0.84]
            check_tan_phi=[0.55,0.60]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[99,728]
            check_z_L=[0.2,1]
            check_z_D=[1,5]
            check_L_D=[3,5]
            
            #PM1
            check_Dr=[0.55,0.92]
            check_tan_phi=[0.70,0.84]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[99,728]
            check_z_L=[0.2,0.8]
            check_z_D=[1,3]
            check_L_D=[3,5]
            
            #PM2
            check_Dr=[0.55,0.92]
            check_tan_phi=[0.70,0.84]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[14,1200]
            check_z_L=[0,1]
            check_z_D=[0,5]
            check_L_D=[3,5]
            
            #PM3
            check_Dr=[0.51,1]
            check_tan_phi=[0.55,0.65]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[14,1200]
            check_z_L=[0.1,1]
            check_z_D=[0.8,5]
            check_L_D=[3,5]
            
            #PM4
            check_Dr=[0.37,0.94]
            check_tan_phi=[0.55,0.65]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[14,1200]
            check_z_L=[0.1,1]
            check_z_D=[0.5,5]
            check_L_D=[3,5]
            
            #PM5
            check_Dr=[0.22,1]
            check_tan_phi=[0.55,0.94]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[14,1200]
            check_z_L=[0,1]
            check_z_D=[0,5]
            check_L_D=[3,5]
            
            #CC1
            check_Dr=[0.38,0.94]
            check_tan_phi=[0.60,0.81]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[14,1200]
            check_z_L=[0.2,1]
            check_z_D=[1,5]
            check_L_D=[3,5]
            
            #CC2
            check_Dr=[0.38,0.94]
            check_tan_phi=[0.60,0.81]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[45,1200]
            check_z_L=[0,1]
            check_z_D=[0,5]
            check_L_D=[3,5]
            
            #CC3
            check_Dr=[0.38,0.94]
            check_tan_phi=[0.60,0.81]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[39,815]
            check_z_L=[0,1]
            check_z_D=[0,5]
            check_L_D=[3,5]
            
            #ALL
            check_Dr=[0.22,1]
            check_tan_phi=[0.53,1.07]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[0,1200]
            check_z_L=[0,1]
            check_z_D=[0,5]
            check_L_D=[3,5]
            
            #list6=['Dr','tan_phi','Su','CSR','CF']
            
            #check_sand=[check_Dr,check_tan_phi,check_CSR,check_CF]
            #check_clay=[check_Su,check_CSR,check_CF]
            
            #check_g=[check_z_L,check_z_D,check_L_D]
            
            #if(Unit[0]=='SS1' or Unit[0]=='H2' or Unit[0]=='PC1' or Unit[0]=='PH1' or Unit[0]=='PM1' or Unit[0]=='PM3' or Unit[0]=='PM4' or Unit[0]=='PM5' or Unit[0]=='CC1'):
            #    soil_type_str="Sand"
            #    check=check_sand
            #else:
            #    soil_type_str="Clay"
            #    check=check_clay
             
            dependant_f_str=array_full_str[7]
            dependant_g_str=array_full_str[9]
            type_f=array_full_str[3]
            type_g=array_full_str[5]
            
            #Parameters check
            if (dependant_f_str=='Dr'):
                check_f=check_Dr
            elif (dependant_f_str=='tan_phi'):
                check_f=check_tan_phi
            elif (dependant_f_str=='Su'):
                check_f=check_Su
            elif (dependant_f_str=='CSR'):
                check_f=check_CSR
            elif (dependant_f_str=='CF'):
                check_f=check_CF
            
            
            if (dependant_g_str=='z_L'):
                check_g=check_z_L
            elif (dependant_g_str=='z_D'):
                check_g=check_z_D
            elif (dependant_g_str=='L_D'):
                check_g=check_L_D
                
            
            PU_brut=[]
            for index_check_f in range(len(check_f)):
                for index_check_g in range(len(check_g)):
                    
                    dependant_g=check_g[index_check_g]
                    dependant_f=check_f[index_check_f]
                    
# =============================================================================
#                     print(dependant_g)
#                     print(dependant_f)
#                     print(type_f)
#                     print(type_g)
#                     print(Level_1[0])
# =============================================================================
                    print(Level_1[0])
                    
                    if type_f=='0' and type_g=='0':
                        Level_1_1=Level_1[0]
                        Level_2_1=Level_1_1
                        Level_3=Level_2_1
                    elif type_f=='0' and type_g=='1':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_2_1=Level_1_1
                        Level_2_2=Level_1_2
                        Level_3=Level_2_1+Level_2_2*dependant_g
                    elif type_f=='0' and type_g=='2':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=Level_1[2]
                        Level_2_1=Level_1_1
                        Level_2_2=Level_1_2
                        Level_2_3=Level_1_3
                        Level_3=Level_2_1+Level_2_2*np.exp(Level_2_3*dependant_g)
                    elif type_f=='1' and type_g=='0':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_2_1=Level_1_1+Level_1_2*dependant_f
                        Level_3=Level_2_1
                    elif type_f=='1' and type_g=='1':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=Level_1[2]
                        Level_1_4=Level_1[3]
                        Level_2_1=Level_1_1+Level_1_2*dependant_f
                        Level_2_2=Level_1_3+Level_1_4*dependant_f
                        Level_3=Level_2_1+Level_2_2*dependant_g
                    elif type_f=='1' and type_g=='2':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=Level_1[2]
                        Level_1_4=Level_1[3]
                        Level_1_5=Level_1[4]
                        Level_1_6=Level_1[5]
                        Level_2_1=Level_1_1+Level_1_2*dependant_f
                        Level_2_2=Level_1_3+Level_1_4*dependant_f
                        Level_2_3=Level_1_5+Level_1_6*dependant_f
                        Level_3=Level_2_1+Level_2_2*np.exp(Level_2_3*dependant_g)
                    elif type_f=='2' and type_g=='0':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=Level_1[2]
                        Level_2_1=Level_1_1+Level_1_2*np.exp(Level_1_3*dependant_f)
                        Level_3=Level_2_1
                    elif type_f=='2' and type_g=='1':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=Level_1[2]
                        Level_1_4=Level_1[3]
                        Level_1_5=Level_1[4]
                        Level_1_6=Level_1[5]
                        Level_2_1=Level_1_1+Level_1_2*np.exp(Level_1_3*dependant_f)
                        Level_2_2=Level_1_4+Level_1_5*np.exp(Level_1_6*dependant_f)
                        Level_3=Level_2_1+Level_2_2*dependant_g
                    elif type_f=='2' and type_g=='2':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=Level_1[2]
                        Level_1_4=Level_1[3]
                        Level_1_5=Level_1[4]
                        Level_1_6=Level_1[5]
                        Level_1_7=Level_1[6]
                        Level_1_8=Level_1[7]
                        Level_1_9=Level_1[8]
                        Level_2_1=Level_1_1+Level_1_2*np.exp(Level_1_3*dependant_f)
                        Level_2_2=Level_1_4+Level_1_5*np.exp(Level_1_6*dependant_f)
                        Level_2_3=Level_1_7+Level_1_8*np.exp(Level_1_9*dependant_f)
                        Level_3=Level_2_1+Level_2_2*np.exp(Level_2_3*dependant_g)
                    
                    PU_brut=PU_brut+[Level_3]
                    print("PU: ",Level_3)
                
            if PU_brut[0]>=0 and PU_brut[1]>=0 and PU_brut[2]>=0 and PU_brut[3]>=0:
                print('Validated - Automatic')
                print(array_full_str)
                print(array_to_show)
                title_path=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                full_path=(my_path+subdirectory+'/'+title_path+'.png')
                
                layout = [ [sg.Text('Best prediction for parameters defined:')],
                           [sg.Text('Are you happy with this prediction?')],
                           #[sg.Text('Source for Files', size=(15, 1)), sg.InputText(), sg.FilesBrowse()],
                           [sg.Button('Yes'), sg.Button('No')],
                           [sg.Image(full_path)]]
                           


                layout = layout 
                window = sg.Window('Final prediction',layout,finalize=True)


                #Read GUI values
                while True: 
                    event, values = window.read()
                    output=''

                    if event == sg.WIN_CLOSED or event == 'Yes':
                        output='Yes'
                        break
                    
                    if event == 'No':
                        output='No'
                        break
                    
                window.close()
                
                if output=='Yes':
                    my_file_kid='boy-kid'
                    full_path_kid=(my_path+'/'+my_file_kid+'.gif')
                    
                    if type_f=='0' and type_g=='0':
                        Level_1_1=Level_1[0]
                        Level_1_2=None
                        Level_1_3=None
                        Level_2_1=None
                        Level_2_2=None
                        Level_2_3=None
                        Level_3_1=None
                        Level_3_2=None
                        Level_3_3=None

                    elif type_f=='0' and type_g=='1':
                        Level_1_1=Level_1[0]
                        Level_1_2=None
                        Level_1_3=None
                        Level_2_1=Level_1[1]
                        Level_2_2=None
                        Level_2_3=None
                        Level_3_1=None
                        Level_3_2=None
                        Level_3_3=None

                    elif type_f=='0' and type_g=='2':
                        Level_1_1=Level_1[0]
                        Level_1_2=None
                        Level_1_3=None
                        Level_2_1=Level_1[1]
                        Level_2_2=None
                        Level_2_3=None
                        Level_3_1=Level_1[2]
                        Level_3_2=None
                        Level_3_3=None

                    elif type_f=='1' and type_g=='0':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=None
                        Level_2_1=None
                        Level_2_2=None
                        Level_2_3=None
                        Level_3_1=None
                        Level_3_2=None
                        Level_3_3=None

                    elif type_f=='1' and type_g=='1':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=None
                        Level_2_1=Level_1[2]
                        Level_2_2=Level_1[3]
                        Level_2_3=None
                        Level_3_1=None
                        Level_3_2=None
                        Level_3_3=None

                    elif type_f=='1' and type_g=='2':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=None
                        Level_2_1=Level_1[2]
                        Level_2_2=Level_1[3]
                        Level_2_3=None
                        Level_3_1=Level_1[4]
                        Level_3_2=Level_1[5]
                        Level_3_3=None

                    elif type_f=='2' and type_g=='0':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=Level_1[2]
                        Level_2_1=None
                        Level_2_2=None
                        Level_2_3=None
                        Level_3_1=None
                        Level_3_2=None
                        Level_3_3=None

                    elif type_f=='2' and type_g=='1':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=Level_1[2]
                        Level_2_1=Level_1[3]
                        Level_2_2=Level_1[4]
                        Level_2_3=Level_1[5]
                        Level_3_1=None
                        Level_3_2=None
                        Level_3_3=None

                    elif type_f=='2' and type_g=='2':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=Level_1[2]
                        Level_2_1=Level_1[3]
                        Level_2_2=Level_1[4]
                        Level_2_3=Level_1[5]
                        Level_3_1=Level_1[6]
                        Level_3_2=Level_1[7]
                        Level_3_3=Level_1[8]

                    
                    layout = [ [sg.Text('Well done!')],
                               [sg.Image(full_path_kid, key='_IMAGE_')],
                               #[sg.Image(full_path)],
                               #[sg.Text('Source for Files', size=(15, 1)), sg.InputText(), sg.FilesBrowse()],
                               [sg.Text(str('Soil unit: '+Unit[0]+' - '+'Parameter: '+ array_full_str[1]))],
                               [sg.Text(str('Variable_1_1: '+str(Level_1_1)))],
                               [sg.Text(str('Variable_1_2: '+str(Level_1_2)))],
                               [sg.Text(str('Variable_1_3: '+str(Level_1_3)))],
                               [sg.Text(str('Variable_2_1: '+str(Level_2_1)))],
                               [sg.Text(str('Variable_2_2: '+str(Level_2_2)))],
                               [sg.Text(str('Variable_2_3: '+str(Level_2_3)))],
                               [sg.Text(str('Variable_3_1: '+str(Level_3_1)))],
                               [sg.Text(str('Variable_3_2: '+str(Level_3_2)))],
                               [sg.Text(str('Variable_3_3: '+str(Level_3_3)))],
                               [sg.Text(str('Type_f: '+type_f+' - '+'Dependant_f: '+ dependant_f_str))],
                               [sg.Text(str('Type_g: '+type_g+' - '+'Dependant_g: '+ dependant_g_str))],
                               [sg.Text('Project: '),sg.Input(k='-IN-'),sg.Text(size=(14,1), key='-OUT-')],
                               [sg.Text('Revision: '),sg.Input(k='-IN-0'),sg.Text(size=(14,1), key='-OUT-0')],
                               [sg.Button('Save Table in DB')]]
                    layout = layout 
                    window = sg.Window('Final prediction',layout,finalize=True)
                    
                    while True: 
                        event, values = window.read(timeout=25)
                        if event in (None, 'Cancel'):
                            break
                        window.Element('_IMAGE_').UpdateAnimation(full_path_kid, time_between_frames=50)
                        
                        if event == sg.WIN_CLOSED :
                            save_table='No'
                            break
                        
                        if event == 'Save Table in DB' :
                            save_table='Yes'
                            break
                        
                        window['-OUT-'].update(values['-IN-'])
                    
                    window.close()
                    
                    if save_table=='Yes':
                        print('Results saved.')
                        Project_name=values['-IN-']
                        Revision_number=values['-IN-0']
                        
                        conn=mysql.connector.connect(user='owdb_user',password='ituotdowdb',host='dklycopilod1',database='owdb')
                        mycursor=conn.cursor()
                        
                        
                        sql=str('INSERT INTO Pisa_Param_Level_1 '+'(Project,Soil_unit,Parameter,Type_f,Dep_f,Type_g,Dep_g,rev,Coe_1_1,Coe_1_2,Coe_1_3,Coe_2_1,Coe_2_2,Coe_2_3,Coe_3_1,Coe_3_2,Coe_3_3) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)')
                        #val = str(sql + "INSERT INTO customers (name, address) VALUES (%s, %s)")
                        val = (str(Project_name), str(Unit[0]), str(array_full_str[1]), type_f, dependant_f_str, type_g, dependant_g_str, Revision_number, Level_1_1, Level_1_2, Level_1_3, Level_2_1, Level_2_2, Level_2_3, Level_3_1, Level_3_2, Level_3_3)
                        
                        mycursor.execute(sql, val)
                        
                        conn.commit()
                        
                        print(mycursor.rowcount, "record inserted.")
                        
                    else:
                        print('Results not saved.')
                        
                    break
                
                    
                    
                else:
                    print('Not validated - User choice')
            else:
                print('Not validated - Automatic')
            
        
        
        #Check automatic
        #Rank the error_table
        #For loop over the ranked table
        #Check if PU >0
        #If yes print final
        #If not go to the next one and test
        
    else :
        print('==================================================================')
        print('Non automatic regression selected!')
        
        k=0
        for j in range(len(df1)):
            #print(j)
            if df1['Unit_name'].values[j]==values[list_toDo[i]*8] and (df1['Cons_Name'].values[j]==values[list_toDo[i]*8+6] or values[list_toDo[i]*8+6]=='ALL') and (df1['Contour_diagram_name'].values[j]==values[list_toDo[i]*8+7] or values[list_toDo[i]*8+7]=='ALL'):
                k=k+1
            else:
                #remove the lines that do not validate: unit / model / contour chosen
                array_regression=np.delete(array_regression,k,0)
    
        #make read dependant f, and type functions
        dependant_f_str=values[list_toDo[i]*8+5]
        type_f=values[list_toDo[i]*8+4]
        type_g=values[list_toDo[i]*8+3]
        param_regression_str=values[list_toDo[i]*8+1]
        
        # make a case if clay/sand + param chosen to define dependant g                  
        print('==================================================================')
        print('Array extracted from mySQL:')             
        print(array_regression)

        
        Unit=array_regression[:,0]
        Model=array_regression[:,1]
        Project=array_regression[:,2]
        Location=array_regression[:,3]
        Rev=array_regression[:,4]
        Layer=array_regression[:,5]
        depth=-array_regression[:,6]
        length=array_regression[:,7]
        diameter=array_regression[:,8]
        
        YU=array_regression[:,9]
        PU=array_regression[:,10]
        KP=array_regression[:,11]
        NP=array_regression[:,12]
        TETA=array_regression[:,13]
        MU=array_regression[:,14]
        KM=array_regression[:,15]
        NM=array_regression[:,16]
        YB=array_regression[:,17]
        HB=array_regression[:,18]
        KB=array_regression[:,19]
        NB=array_regression[:,20]
        TETAMB=array_regression[:,21]
        MB=array_regression[:,22]
        KMB=array_regression[:,23]
        NMB=array_regression[:,24]
        
        ASR=array_regression[:,25]
        Dr=array_regression[:,26]
        phi=array_regression[:,27]
        phi_rad=[0]*(len(phi))
        for i in range(len(phi)):
            phi_rad[i]=np.radians(phi[i])
        Su=array_regression[:,28]  
        G0=array_regression[:,29]   
        SigmaV=array_regression[:,30]   
        CSR=array_regression[:,31]   
        Neq=array_regression[:,32]   
        CF_OCR=array_regression[:,33] 
        CF_Dr=array_regression[:,34] 
        Batch=array_regression[:,35] 
        
        z_L=depth/length
        z_D=depth/diameter
        L_D=length/diameter
        Dr=Dr*0.01
        tan_phi=np.tan(phi_rad)
        CF=CF_OCR*CF_Dr
    
                
        input_data={}
        input_data['Unit']=Unit
        input_data['Model']=Model
        input_data['Project']=Project
        input_data['Location']=Location
        input_data['Rev']=Rev
        input_data['Layer']=Layer
        input_data['depth']=depth
        input_data['length']=length
        input_data['diameter']=diameter
        
        input_data['Yu']=YU
        input_data['Pu']=PU
        input_data['Kp']=KP
        input_data['np']=NP
        input_data['Teta']=TETA
        input_data['Mu']=MU
        input_data['Km']=KM
        input_data['nm']=NM
        input_data['Yb']=YB
        input_data['Hb']=HB
        input_data['Kb']=KB
        input_data['nb']=NB
        input_data['TetaMb']=TETAMB
        input_data['Mb']=MB
        input_data['kMb']=KMB
        input_data['nMb']=NMB
        
        list2=['Yu','Pu','Kp','np','Teta','Mu','Km','nm','Yb','Hb','Kb','nb','TetaMb','Mb','kMb','nMb']
        
        input_data['ASR']=ASR
        input_data['Dr']=Dr
        input_data['phi']=phi
        input_data['phi_rad']=phi_rad
        input_data['Su']=Su
        input_data['G0']=G0
        input_data['SigmaV']=SigmaV
        input_data['CSR']=CSR
        input_data['Neq']=Neq
        input_data['CF_OCR']=CF_OCR
        input_data['CF_Dr']=CF_Dr
        input_data['Batch']=Batch
        
        input_data['CF']= CF
        input_data['z_L']=z_L
        input_data['z_D']=z_D
        input_data['L_D']=L_D
        input_data['tan_phi']=tan_phi
        
        input_data['type_function_f']=type_f
        input_data['type_function_g']=type_g
        
        #print(Unit[0])
        #print(param_regression_str)
        
        if(Unit[0]=='SS1' or Unit[0]=='H2' or Unit[0]=='PC1' or Unit[0]=='PH1' or Unit[0]=='PM1' or Unit[0]=='PM3' or Unit[0]=='PM4' or Unit[0]=='PM5' or Unit[0]=='CC1'):
            if(param_regression_str=='Yb' or param_regression_str=='Hb' or param_regression_str=='Kb' or param_regression_str=='nb'):
                input_data['dependant_g']=L_D
                dependant_g_str='L_D'
            else:
                input_data['dependant_g']=z_L
                dependant_g_str='z_L'
        else:
            if(param_regression_str=='Yu' or param_regression_str=='Pu' or param_regression_str=='Kp' or param_regression_str=='np'):
                input_data['dependant_g']=z_D
                dependant_g_str='z_D'
            elif (param_regression_str=='Teta' or param_regression_str=='Mu' or param_regression_str=='Km' or param_regression_str=='nm'):
                input_data['dependant_g']=z_L
                dependant_g_str='z_L'
            else:
                input_data['dependant_g']=L_D
                dependant_g_str='L_D'
            
        print(dependant_g_str)
        input_data['dependant_f']=input_data[dependant_f_str]
        param_regression=input_data[param_regression_str]
        
        print('==================================================================')
        #print(range(len(PISA)))
        #print(type(PISA))
        try:
            poptM, pcovM = curve_fit(g2,input_data, param_regression)
            print(poptM)
            #print(variable_total)
            predict_param=g2(input_data,poptM[0],poptM[1],poptM[2],poptM[3],poptM[4],poptM[5],poptM[6],poptM[7],poptM[8]) 
            
            index_constant=0
            if(type_g=='0' and type_f=='0'):
                constant=predict_param
                predict_param=[]
                for index_constant in range(len(param_regression)):
                    predict_param=predict_param+[constant]
           
            print('Prediction: ', predict_param)
            
            depth_profile=np.amax(depth)
            calc_y=round(depth_profile/5)+1
                  
            M_profile=np.amax(param_regression)
            calc_yM=round(M_profile/100)+1
                  
            fig = plt.figure(figsize=(10, 10))
            ax1 = fig.add_subplot(111)
                  #ax1.grid()
                  
            line1=plt.scatter(param_regression,depth)
            line1.set_label("Real values")
                  
            line6=plt.scatter(predict_param,depth,color='black')
            line6.set_label('Predicted values')
              
            ax1.set_xlabel(param_regression_str)
                  # ax1.set_xlim(0,max(Gmax))
            min_depth=0
            max_depth=calc_y*5
            plt.ylim([min_depth,max_depth])
            plt.xlim([0,calc_yM*M_profile*2])
            plt.gca().invert_yaxis()
            ax1.set_ylabel('Depth [m]')
                  
            ax1.legend()
            
            if(type_g=='0'):
                if(type_f=='0'):
                    plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+") )")
                    title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                elif(type_f=='1'):
                    plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+ "*"+dependant_f_str +") )")
                    title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                else:
                    plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+"*exp("+str(round(poptM[2],2))+"*"+dependant_f_str +") )")
                    title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
            elif(type_g=='1'):
                if(type_f=='0'):
                    plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[1],2))+"))"+" *"+dependant_g_str+" )")
                    title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                elif(type_f=='1'):
                    plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+ "*"+dependant_f_str +") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[2],2))+"+"+str(round(poptM[3],2))+ "*"+dependant_f_str+"))"+" *"+dependant_g_str+" )")
                    title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                else:
                    plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+"*exp("+str(round(poptM[2],2))+ "*"+dependant_f_str +") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[3],2))+"+"+str(round(poptM[4],2))+"*exp("+str(round(poptM[5],2))+ "*"+dependant_f_str+"))"+" *"+dependant_g_str+" )")
                    title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
            else:
                if(type_f=='0'):
                     plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[1],2))+"))"+"*exp("+param_regression_str+"_3= ("+str(round(poptM[2],2))+"))"+" *"+dependant_g_str+") )")
                     title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                elif(type_f=='1'):
                     plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+ "*"+dependant_f_str +") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[2],2))+"+"+str(round(poptM[3],2))+ "*"+dependant_f_str+"))"+"*exp("+param_regression_str+"_3= ("+str(round(poptM[4],2))+"+"+str(round(poptM[5],2))+"*"+dependant_f_str+")"+" *"+dependant_g_str+") )")
                     title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                else:
                     plt.title(Unit[0]+": "+param_regression_str+"= ( "+param_regression_str+"_1= ("+str(round(poptM[0],2))+"+"+str(round(poptM[1],2))+"*exp("+str(round(poptM[2],2))+ "*"+dependant_f_str +") "+"+ ("+param_regression_str+"_2= ("+str(round(poptM[3],2))+"+"+str(round(poptM[4],2))+"*exp("+str(round(poptM[5],2))+ "*"+dependant_f_str+"))"+"*exp("+param_regression_str+"_3= ("+str(round(poptM[6],2))+"+"+str(round(poptM[7],2))+"*exp("+str(round(poptM[8],2))+"*"+dependant_f_str+"))"" *"+dependant_g_str+") )")
                     title_plot=str(Unit[0]+"_"+param_regression_str+"_f_"+type_f+"_g_"+type_g+"_dep_f_"+dependant_f_str+"_dep_g_"+dependant_g_str)
                     
            #variable_total= [0.3667, 25.89, 0.3375, -8.9, 1, 1]     

            my_path = os.path.dirname(__file__)
            my_file = title_plot
            subdirectory=('\Plots_'+values['-IN-']+'_'+values['-IN-0'])
            full_path=(my_path+subdirectory+'/'+my_file+'.png')
            print(my_path)
            print(my_file)
            print(subdirectory)
            if not os.path.isdir(my_path+subdirectory):
                os.makedirs(my_path+subdirectory)
    
            fig.savefig(full_path)
            
            title_excel_original='\Template.xlsx'
            title_excel_target='\Regression_parameters.xlsx'
            
            original=(my_path+title_excel_original)
            target=(my_path+subdirectory+title_excel_target)
            
            if not os.path.isfile(my_path+subdirectory+title_excel_target):
                shutil.copyfile(original, target)
            
            my_file_excel=str(Unit[0]+"_"+param_regression_str+"_"+type_f+"_"+type_g+"_"+dependant_f_str+"_"+dependant_g_str)
            
            workbook = openpyxl.load_workbook(target)
            if my_file_excel in workbook.sheetnames:
                workbook.remove_sheet(workbook.get_sheet_by_name(my_file_excel))
                workbook.create_sheet(my_file_excel)
            else: 
                workbook.create_sheet(my_file_excel)
                
            sheet1=workbook['Template']
            sheet2=workbook[my_file_excel]
            
            maxr = sheet1.max_row
            maxc = sheet1.max_column

            for r in range (1, maxr + 1):
                for c in range (1, maxc + 1):
                    sheet2.cell(row=r,column=c).value = sheet1.cell(row=r,column=c).value
            
            index_row=0
            index_column=0
            
            max_array_r=len(array_regression)
            max_array_c=36
            
            for index_row in range(max_array_r):
                for index_column in range(max_array_c):
                    sheet2.cell(row=index_row+3, column=index_column+1).value=array_regression[index_row][index_column]
            
            index_variable=0
            
            for index_variable in range(len(poptM)):
                sheet2.cell(row=3, column=index_variable+44).value=poptM[index_variable]
            
            sheet2.cell(row=3, column=54).value=dependant_f_str
            sheet2.cell(row=3, column=55).value=dependant_g_str
            sheet2.cell(row=3, column=57).value=type_f
            sheet2.cell(row=3, column=58).value=type_g
            
            sheet2.cell(row=3, column=38).value=param_regression_str
            
            
            
            
            error=Error_function(param_regression,input_data,poptM[0],poptM[1],poptM[2],poptM[3],poptM[4],poptM[5],poptM[6],poptM[7],poptM[8])
            print('==================================================================')
            print('Error_predict: ', error) 
            
            sheet2.cell(row=3, column=60).value=error
            workbook.save(target)
            workbook.close()
            
            #Parameters level 1
            Level_1=poptM
            #dependant_g=input_data['dependant_g']
            #dependant_f=input_data['dependant_f']
            

            
            
            check_Dr=[0.22,1]
            check_tan_phi=[0.53,1.07]
            check_CSR=[0.22,1.4]
            check_CF=[0.5,4.6]
            check_Su=[0,1200]
            check_z_L=[0,1]
            check_z_D=[0,5]
            check_L_D=[3,5]
            
            #list6=['Dr','tan_phi','Su','CSR','CF']
            
            #check_sand=[check_Dr,check_tan_phi,check_CSR,check_CF]
            #check_clay=[check_Su,check_CSR,check_CF]
            
            #check_g=[check_z_L,check_z_D,check_L_D]
            
            #if(Unit[0]=='SS1' or Unit[0]=='H2' or Unit[0]=='PC1' or Unit[0]=='PH1' or Unit[0]=='PM1' or Unit[0]=='PM3' or Unit[0]=='PM4' or Unit[0]=='PM5' or Unit[0]=='CC1'):
            #    soil_type_str="Sand"
            #    check=check_sand
            #else:
            #    soil_type_str="Clay"
            #    check=check_clay
                
            
            #Parameters check
            if (dependant_f_str=='Dr'):
                check_f=check_Dr
            elif (dependant_f_str=='tan_phi'):
                check_f=check_tan_phi
            elif (dependant_f_str=='Su'):
                check_f=check_Su
            elif (dependant_f_str=='CSR'):
                check_f=check_CSR
            elif (dependant_f_str=='CF'):
                check_f=check_CF
            
            
            if (dependant_g_str=='z_L'):
                check_g=check_z_L
            elif (dependant_g_str=='z_D'):
                check_g=check_z_D
            elif (dependant_g_str=='L_D'):
                check_g=check_L_D
                
            
            
            for index_check_f in range(len(check_f)):
                for index_check_g in range(len(check_g)):
                    
                    dependant_g=check_g[index_check_g]
                    dependant_f=check_f[index_check_f]
                    
                    if type_f=='0' and type_g=='0':
                        Level_1_1=Level_1[0]
                        Level_2_1=Level_1_1
                        Level_3=Level_2_1
                    elif type_f=='0' and type_g=='1':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_2_1=Level_1_1
                        Level_2_2=Level_1_2
                        Level_3=Level_2_1+Level_2_2*dependant_g
                    elif type_f=='0' and type_g=='2':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=Level_1[2]
                        Level_2_1=Level_1_1
                        Level_2_2=Level_1_2
                        Level_2_3=Level_1_3
                        Level_3=Level_2_1+Level_2_2*np.exp(Level_2_3*dependant_g)
                    elif type_f=='1' and type_g=='0':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_2_1=Level_1_1+Level_1_2*dependant_f
                        Level_3=Level_2_1
                    elif type_f=='1' and type_g=='1':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=Level_1[2]
                        Level_1_4=Level_1[3]
                        Level_2_1=Level_1_1+Level_1_2*dependant_f
                        Level_2_2=Level_1_3+Level_1_4*dependant_f
                        Level_3=Level_2_1+Level_2_2*dependant_g
                    elif type_f=='1' and type_g=='2':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=Level_1[2]
                        Level_1_4=Level_1[3]
                        Level_1_5=Level_1[4]
                        Level_1_6=Level_1[5]
                        Level_2_1=Level_1_1+Level_1_2*dependant_f
                        Level_2_2=Level_1_3+Level_1_4*dependant_f
                        Level_2_3=Level_1_5+Level_1_6*dependant_f
                        Level_3=Level_2_1+Level_2_2*np.exp(Level_2_3*dependant_g)
                    elif type_f=='2' and type_g=='0':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=Level_1[2]
                        Level_2_1=Level_1_1+Level_1_2*np.exp(Level_1_3*dependant_f)
                        Level_3=Level_2_1
                    elif type_f=='2' and type_g=='1':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=Level_1[2]
                        Level_1_4=Level_1[3]
                        Level_1_5=Level_1[4]
                        Level_1_6=Level_1[5]
                        Level_2_1=Level_1_1+Level_1_2*np.exp(Level_1_3*dependant_f)
                        Level_2_2=Level_1_4+Level_1_5*np.exp(Level_1_6*dependant_f)
                        Level_3=Level_2_1+Level_2_2*dependant_g
                    elif type_f=='2' and type_g=='2':
                        Level_1_1=Level_1[0]
                        Level_1_2=Level_1[1]
                        Level_1_3=Level_1[2]
                        Level_1_4=Level_1[3]
                        Level_1_5=Level_1[4]
                        Level_1_6=Level_1[5]
                        Level_1_7=Level_1[6]
                        Level_1_8=Level_1[7]
                        Level_1_9=Level_1[8]
                        Level_2_1=Level_1_1+Level_1_2*np.exp(Level_1_3*dependant_f)
                        Level_2_2=Level_1_4+Level_1_5*np.exp(Level_1_6*dependant_f)
                        Level_2_3=Level_1_7+Level_1_8*np.exp(Level_1_9*dependant_f)
                        Level_3=Level_2_1+Level_2_2*np.exp(Level_2_3*dependant_g)
        
                    print("PU: ",Level_3)
                    
                    if(Level_3 >=0):
                        print('Validated')
                    else:
                        print('Not validated')
                    
        except:
            print('Maximum iteration reached for CurveFit function')          
        
            pass

